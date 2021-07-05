from config.file_path import *
from openpyxl import load_workbook
import yaml


def read_yaml(path):
    with open(path,mode='r',encoding='utf-8') as f:
        data=yaml.safe_load(f)
    return data

def write_excel(path,data):
    wb = load_workbook(path)
    for sheet in wb:
        sheet.append(data)
    wb.save(path)

def read_excel(path,line_num):
    wb=load_workbook(path)
    sheet=wb['Sheet1']
    datalist=[]
    for line in list(sheet.rows)[1:]:
        datalist.append(line[line_num].value)
    return datalist


def read_picture(path):
    with open(path, 'rb') as f:
        file=f.read()
        return file
